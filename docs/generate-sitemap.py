#!/usr/bin/env python3
"""
Generate the Atelier Rusalka sitemap tracker (.xlsx) from sitemap.json.

Usage:
    python3 docs/generate-sitemap.py                  # writes to docs/Atelier-Rusalka-Sitemap.xlsx
    python3 docs/generate-sitemap.py --out OTHER.xlsx # writes elsewhere

The JSON file is the source of truth for STRUCTURE (sections, pages, tasks).
Once the .xlsx is generated, edit STATUS / OWNER / NOTES directly in Excel —
re-running the script will OVERWRITE the file, so do not re-run on top of
manual edits unless you've added new pages/tasks to the JSON and want to
discard prior status edits.

Modular: to add a section/page/task, edit sitemap.json and re-run.
"""

from __future__ import annotations

import argparse
import json
import os
import sys
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import (
        Alignment,
        Border,
        Font,
        PatternFill,
        Side,
    )
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.formatting.rule import CellIsRule
except ImportError:
    sys.stderr.write(
        "openpyxl is required. Install with:\n"
        "  pip3 install --user openpyxl\n"
    )
    sys.exit(1)


# ---------- Constants ----------------------------------------------------- #

HERE = Path(__file__).resolve().parent
DEFAULT_JSON = HERE / "sitemap.json"
DEFAULT_XLSX = HERE / "Atelier-Rusalka-Sitemap.xlsx"

# Visual tokens — keep them in one place so the look is easy to tune.
COLORS = {
    "section_bg": "1F1D1A",        # warm ink
    "section_fg": "F8F2E6",        # cream
    "page_bg":    "EBE2CF",        # soft tint of bg
    "page_fg":    "1F1D1A",
    "header_bg":  "3A352D",
    "header_fg":  "F8F2E6",
    "status_done":        "D7F4E0",
    "status_in_progress": "FFF1CC",
    "status_todo":        "F2EFE7",
    "status_blocked":     "FAD7D7",
    "status_skipped":     "EAEAEA",
    "border":             "DDD8CF",
}

COLUMNS = [
    ("Section",     14),
    ("Page",        24),
    ("Path",        28),
    ("Task",        58),
    ("Status",      14),
    ("Priority",    14),
    ("Owner",       14),
    ("Notes",       46),
    ("Updated",     12),
]


# ---------- Helpers ------------------------------------------------------- #

def cell_border(color: str = COLORS["border"]) -> Border:
    side = Side(style="thin", color=color)
    return Border(left=side, right=side, top=side, bottom=side)


def fill(color: str) -> PatternFill:
    return PatternFill(start_color=color, end_color=color, fill_type="solid")


def status_fill(status: str) -> PatternFill | None:
    key = (status or "").strip().lower()
    table = {
        "done":        COLORS["status_done"],
        "in progress": COLORS["status_in_progress"],
        "todo":        COLORS["status_todo"],
        "blocked":     COLORS["status_blocked"],
        "skipped":     COLORS["status_skipped"],
        "n/a":         COLORS["status_skipped"],
    }
    color = table.get(key)
    return fill(color) if color else None


# ---------- Main build ---------------------------------------------------- #

def build_workbook(data: dict) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "Sitemap"

    # --- Header row ---------------------------------------------------- #
    header_fill = fill(COLORS["header_bg"])
    header_font = Font(
        name="Helvetica",
        size=10,
        bold=True,
        color=COLORS["header_fg"],
    )
    for col_idx, (label, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(
            horizontal="left", vertical="center", wrap_text=False,
        )
        cell.border = cell_border(COLORS["header_bg"])
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 26
    ws.freeze_panes = "A2"

    # --- Section / page / task rows ----------------------------------- #
    row = 2
    section_fill_obj = fill(COLORS["section_bg"])
    section_font = Font(name="Helvetica", size=11, bold=True, color=COLORS["section_fg"])
    page_fill_obj = fill(COLORS["page_bg"])
    page_font = Font(name="Helvetica", size=10, bold=True, color=COLORS["page_fg"])
    body_font = Font(name="Helvetica", size=10)
    align_top_wrap = Alignment(horizontal="left", vertical="top", wrap_text=True)

    metadata = data.get("metadata", {})
    status_values = metadata.get(
        "status_values", ["Todo", "In progress", "Done", "Blocked", "Skipped", "N/A"],
    )
    priority_values = metadata.get(
        "priority_values", ["P0 critical", "P1 high", "P2 medium", "P3 low"],
    )
    base_url = metadata.get("base_url", "").rstrip("/")
    link_font = Font(
        name="Helvetica", size=10, bold=True,
        color="0563C1", underline="single",
    )

    # Data validation — status column (E) and priority column (F).
    # openpyxl quotes the comma-separated list as a literal source.
    status_dv = DataValidation(
        type="list",
        formula1='"%s"' % ",".join(status_values),
        allow_blank=True,
        showDropDown=False,
    )
    status_dv.error = "Pick a status from the dropdown"
    status_dv.errorTitle = "Invalid status"
    ws.add_data_validation(status_dv)

    priority_dv = DataValidation(
        type="list",
        formula1='"%s"' % ",".join(priority_values),
        allow_blank=True,
        showDropDown=False,
    )
    ws.add_data_validation(priority_dv)

    sections = data.get("sections", [])
    for section in sections:
        section_name = section.get("name", "")

        # Section header row — merged across all columns
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(COLUMNS))
        cell = ws.cell(row=row, column=1, value=section_name.upper())
        cell.fill = section_fill_obj
        cell.font = section_font
        cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[row].height = 22
        row += 1

        for page in section.get("pages", []):
            label = page.get("label", "")
            path = page.get("path", "")
            description = page.get("description", "")

            # Page summary row — merged "Task" cell carries description
            ws.cell(row=row, column=1, value=section_name).fill = page_fill_obj
            ws.cell(row=row, column=2, value=label).fill = page_fill_obj
            path_cell = ws.cell(row=row, column=3, value=path)
            path_cell.fill = page_fill_obj
            # Make in-app paths clickable. Skip non-routes like
            # "fulfillment-studio (separate dashboard)".
            if base_url and path.startswith("/"):
                path_cell.hyperlink = f"{base_url}{path}"
                path_cell.font = link_font
            ws.cell(row=row, column=4, value=description).fill = page_fill_obj
            ws.cell(row=row, column=5, value="").fill = page_fill_obj
            ws.cell(row=row, column=6, value="").fill = page_fill_obj
            ws.cell(row=row, column=7, value="").fill = page_fill_obj
            ws.cell(row=row, column=8, value="").fill = page_fill_obj
            ws.cell(row=row, column=9, value="").fill = page_fill_obj
            for col_idx in range(1, len(COLUMNS) + 1):
                cell = ws.cell(row=row, column=col_idx)
                if col_idx == 3 and cell.hyperlink is not None:
                    # Keep the link font we already set above.
                    pass
                elif col_idx in (1, 2):
                    cell.font = page_font
                else:
                    cell.font = body_font
                cell.alignment = align_top_wrap
                cell.border = cell_border()
            ws.row_dimensions[row].height = 20
            row += 1

            for task in page.get("tasks", []):
                values = [
                    section_name,
                    label,
                    path,
                    task.get("name", ""),
                    task.get("status", "Todo"),
                    task.get("priority", ""),
                    task.get("owner", ""),
                    task.get("notes", ""),
                    "",  # Updated column — populated by user
                ]
                for col_idx, val in enumerate(values, start=1):
                    cell = ws.cell(row=row, column=col_idx, value=val)
                    cell.font = body_font
                    cell.alignment = align_top_wrap
                    cell.border = cell_border()

                status_cell = ws.cell(row=row, column=5)
                sf = status_fill(task.get("status", "Todo"))
                if sf:
                    status_cell.fill = sf

                # Attach dropdowns
                status_dv.add(status_cell)
                priority_dv.add(ws.cell(row=row, column=6))

                ws.row_dimensions[row].height = 18
                row += 1

    # --- Conditional formatting (status colour follows user edits) ---- #
    # Range covers all data rows up to row-1 (we just incremented past).
    if row > 2:
        status_range = f"E2:E{row - 1}"
        ws.conditional_formatting.add(
            status_range,
            CellIsRule(
                operator="equal",
                formula=['"Done"'],
                fill=fill(COLORS["status_done"]),
            ),
        )
        ws.conditional_formatting.add(
            status_range,
            CellIsRule(
                operator="equal",
                formula=['"In progress"'],
                fill=fill(COLORS["status_in_progress"]),
            ),
        )
        ws.conditional_formatting.add(
            status_range,
            CellIsRule(
                operator="equal",
                formula=['"Todo"'],
                fill=fill(COLORS["status_todo"]),
            ),
        )
        ws.conditional_formatting.add(
            status_range,
            CellIsRule(
                operator="equal",
                formula=['"Blocked"'],
                fill=fill(COLORS["status_blocked"]),
            ),
        )
        ws.conditional_formatting.add(
            status_range,
            CellIsRule(
                operator="equal",
                formula=['"Skipped"'],
                fill=fill(COLORS["status_skipped"]),
            ),
        )

    return wb


# ---------- Entry point --------------------------------------------------- #

def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--json", type=Path, default=DEFAULT_JSON,
        help=f"Source JSON (default: {DEFAULT_JSON.name})",
    )
    parser.add_argument(
        "--out", type=Path, default=DEFAULT_XLSX,
        help=f"Output XLSX (default: {DEFAULT_XLSX.name})",
    )
    args = parser.parse_args()

    if not args.json.exists():
        sys.stderr.write(f"Missing source: {args.json}\n")
        return 1

    with args.json.open("r", encoding="utf-8") as f:
        data = json.load(f)

    wb = build_workbook(data)
    args.out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(args.out)

    size = os.path.getsize(args.out)
    n_sections = len(data.get("sections", []))
    n_tasks = sum(
        len(p.get("tasks", []))
        for s in data.get("sections", [])
        for p in s.get("pages", [])
    )
    print(
        f"✓ wrote {args.out}  ({size:,} bytes, "
        f"{n_sections} sections, {n_tasks} tasks)",
    )
    return 0


if __name__ == "__main__":
    sys.exit(main())
